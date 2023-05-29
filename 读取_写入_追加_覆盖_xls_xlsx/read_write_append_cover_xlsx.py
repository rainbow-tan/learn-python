import os.path
from typing import List

import openpyxl
from openpyxl.cell import Cell
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet


def read_xlsx(filename: str, sheet_name: str) -> List[list]:
    filename = os.path.abspath(filename)
    assert os.path.isfile(filename), f'{filename} is not file'
    assert filename.lower().endswith('.xlsx'), f'不是.xlsx文件:{filename}'
    wb: Workbook = openpyxl.load_workbook(filename)
    assert sheet_name in wb.sheetnames, "sheet name error"
    ws: Worksheet = wb[sheet_name]
    all_data = []
    for row in ws.rows:
        data_row = []
        for data in row:
            cell: Cell = data
            data_row.append(cell.value)
        all_data.append(data_row)
    print("all data".center(50, '*'))
    for one in all_data:
        print(one)
    return all_data


def write_xlsx(filename: str, sheet_name: str, data: List[list]):
    filename = os.path.abspath(filename)
    assert filename.lower().endswith('.xlsx'), f'不是.xlsx文件:{filename}'
    if not os.path.exists(os.path.dirname(filename)):
        os.makedirs(os.path.dirname(filename))
    wb: Workbook = openpyxl.Workbook()
    sheet: Worksheet = wb.active
    sheet.title = sheet_name

    for row in data:
        sheet.append(row)
    wb.save(filename)
    print(f"保存到:{filename}")


def append_xlsx(filename: str, sheet_name: str, data: List[list]):
    filename = os.path.abspath(filename)
    assert os.path.isfile(filename), f'{filename} is not file'
    assert filename.lower().endswith('.xlsx'), f'不是.xlsx文件:{filename}'
    wb: Workbook = openpyxl.load_workbook(filename)
    assert sheet_name in wb.sheetnames, "sheet name error"
    ws: Worksheet = wb[sheet_name]
    for row in data:
        ws.append(row)
    wb.save('append.xlsx')


def cover_xlsx(filename: str, sheet_name: str, data: List[list]):
    filename = os.path.abspath(filename)
    assert os.path.isfile(filename), f'{filename} is not file'
    assert filename.lower().endswith('.xlsx'), f'不是.xlsx文件:{filename}'
    wb: Workbook = openpyxl.load_workbook(filename)
    assert sheet_name in wb.sheetnames, "sheet name error"
    ws: Worksheet = wb[sheet_name]
    for row_x, row in enumerate(data):
        for clo_x, clo in enumerate(row):
            ws.cell(row_x + 1, clo_x + 1, clo)
    wb.save('over.xlsx')


def main():
    filename = "src.xlsx"
    sheet_name = "Sheet1"
    all_data = read_xlsx(filename, sheet_name)
    write_xlsx('write.xlsx', sheet_name, all_data)
    data = [
        ['a', 'b', 'c', 'd'],
        ['A', 'B', 'C', 'D']
    ]
    append_xlsx(filename, sheet_name, data)
    cover_xlsx(filename, sheet_name, data)


if __name__ == '__main__':
    main()
